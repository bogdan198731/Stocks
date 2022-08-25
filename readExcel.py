from openpyxl import Workbook
from openpyxl import load_workbook
from produs import *
from fileWork import append_to_file

def readExcel(filename, listaProduse):
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    data = ("Nume fisier date" + filename)
    append_to_file("Log.txt", data)
    rand = 2
    for row in sheet.iter_rows(values_only=True):
        column = 2
        nume = str(sheet.cell(rand, column).value)
#        print("Nume = " + nume)
        column += 1
#        print (nume + " " + str(sheet.cell(rand, column).value))
        pret = str(sheet.cell(rand, column).value)
#        print("Pret = " + pret)
        column += 1
        pret_open = str(sheet.cell(rand, column).value)
#        print("pret_open = " + pret_open)
        column += 1
        pret_close = str(sheet.cell(rand, column).value)
#        print("pret_open = " + pret_close)
        prod: produs = produs(nume, pret, pret_open, pret_close)
#        prod.afisareProdus()
        bProd = prod.cautare_produs(prod.nume, listaProduse, pret)
    #    print(" bProd " + str(bProd))
        if not bProd:
            listaProduse.append(prod)
        rand += 1

    print("Nr de produse analizate : " + str(rand))
    
def produs_to_arhiva(lista_prod,filename ):
    workbook = Workbook()
    sheet = workbook.active
    sheet.row = 1
    sheet.column = 1
    for prod in lista_prod:
        sheet.column = 1
        sheet.cell(sheet.row, sheet.column).value = prod.nume
        sheet.column += 1
        sheet.cell(sheet.row, sheet.column).value = prod.tendinta
        sheet.column += 1
        sheet.cell(sheet.row, sheet.column).value = prod.lastPret
        sheet.column += 1
        i=0
        for pret in prod.listaPret:
            i = i+1
            sheet.cell(sheet.row, sheet.column).value = float(pret)
            sheet.column += 1
#            print('i = ' + str(i))
        sheet.row += 1
    workbook.save(filename=filename)
    print('end to excel date')

def readArhiva(filename, listaProduse):
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    data = "Nume fisier arhiva" + filename
    append_to_file("Log.txt", data)
    rand = 1
    for row in sheet.iter_rows(values_only=True):
        column = 1
        nume = str(sheet.cell(rand, column).value)
#        print("Nume = " + nume)
        column += 1
        tendinta = str(sheet.cell(rand, column).value)
        column += 1
        pret = str(sheet.cell(rand, column).value)
        column += 1
#        print("Pret = " + pret)
        prod = produs(nume, pret, 0, 0)
        for i in range(1, prod.MAXNR):
        #    print(i)
            if sheet.cell(rand, column).value == None:
                break
            else:
            #    print("de ce "+ str(sheet.cell(rand, column).value))
                prod.modifPret(sheet.cell(rand, column).value,False)
                prod.tendinta = 0;
                column += 1
        prod.tendinta = int(tendinta);
#        print(" tendinta intare = " + str(prod.tendinta))
        listaProduse.append(prod)
        rand += 1

        